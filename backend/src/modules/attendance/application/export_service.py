"""Service for exporting attendance data to Excel.

Uses openpyxl to generate monthly attendance reports in .xlsx format.
"""

from __future__ import annotations

import calendar
import io
from datetime import date
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.modules.attendance.domain.entities import AttendanceRecord


# Status display mapping (Vietnamese)
STATUS_LABELS = {
    "present": "Có mặt",
    "late": "Muộn",
    "early_leave": "Về sớm",
    "absent": "Vắng",
    "on_leave": "Nghỉ phép",
    "holiday": "Ngày lễ",
}

# Status colors for cells
STATUS_COLORS = {
    "present": "C6EFCE",
    "late": "FFEB9C",
    "early_leave": "FFEB9C",
    "absent": "FFC7CE",
    "on_leave": "BDD7EE",
    "holiday": "D9E2F3",
}


class ExportService:
    """Generates Excel exports for attendance data."""

    def generate_monthly_excel(
        self,
        records: list[AttendanceRecord],
        employee_name: str,
        year: int,
        month: int,
    ) -> io.BytesIO:
        """Generate a monthly attendance Excel report.

        Args:
            records: List of attendance records for the month.
            employee_name: Display name of the employee.
            year: Report year.
            month: Report month.

        Returns:
            BytesIO buffer containing the .xlsx file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"Chấm công T{month}/{year}"

        # Styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"BẢNG CHẤM CÔNG THÁNG {month}/{year}"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        # Employee info
        ws["A3"] = "Nhân viên:"
        ws["B3"] = employee_name
        ws["B3"].font = Font(bold=True)

        ws["A4"] = "Tháng:"
        ws["B4"] = f"{month}/{year}"

        # Summary section
        records_by_date = {r.work_date: r for r in records}
        _, last_day = calendar.monthrange(year, month)

        present_count = sum(1 for r in records if r.status in ("present", "late", "early_leave"))
        late_count = sum(1 for r in records if r.status == "late")
        absent_count = sum(1 for r in records if r.status == "absent")
        leave_count = sum(1 for r in records if r.status == "on_leave")
        total_hours = sum(float(r.work_hours or 0) for r in records)
        total_ot = sum(float(r.overtime_hours or 0) for r in records)

        ws["A6"] = "Tổng kết:"
        ws["A6"].font = header_font
        ws["A7"] = f"Ngày có mặt: {present_count}"
        ws["A8"] = f"Ngày muộn: {late_count}"
        ws["A9"] = f"Ngày vắng: {absent_count}"
        ws["A10"] = f"Ngày nghỉ phép: {leave_count}"
        ws["C7"] = f"Tổng giờ làm: {total_hours:.1f}h"
        ws["C8"] = f"Tổng giờ OT: {total_ot:.1f}h"

        # Table header
        row_start = 12
        headers = ["Ngày", "Thứ", "Check-in", "Check-out", "Giờ làm", "OT", "Trạng thái"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Day names in Vietnamese
        day_names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"]

        # Data rows
        for day in range(1, last_day + 1):
            row = row_start + day
            current_date = date(year, month, day)
            weekday = current_date.weekday()
            record = records_by_date.get(current_date)

            ws.cell(row=row, column=1, value=f"{day:02d}/{month:02d}").border = thin_border
            ws.cell(row=row, column=2, value=day_names[weekday]).border = thin_border

            if record:
                check_in_str = record.check_in.strftime("%H:%M") if record.check_in else "—"
                check_out_str = record.check_out.strftime("%H:%M") if record.check_out else "—"
                work_hours_str = f"{float(record.work_hours):.1f}" if record.work_hours else "—"
                ot_str = f"{float(record.overtime_hours):.1f}" if record.overtime_hours else "0"
                status_label = STATUS_LABELS.get(record.status, record.status)

                ws.cell(row=row, column=3, value=check_in_str).border = thin_border
                ws.cell(row=row, column=4, value=check_out_str).border = thin_border
                ws.cell(row=row, column=5, value=work_hours_str).border = thin_border
                ws.cell(row=row, column=6, value=ot_str).border = thin_border

                status_cell = ws.cell(row=row, column=7, value=status_label)
                status_cell.border = thin_border
                color = STATUS_COLORS.get(record.status, "FFFFFF")
                status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            else:
                # Weekend or no record
                if weekday >= 5:  # Saturday/Sunday
                    for col_idx in range(3, 8):
                        cell = ws.cell(row=row, column=col_idx, value="—")
                        cell.border = thin_border
                        cell.fill = PatternFill(
                            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                        )
                else:
                    for col_idx in range(3, 8):
                        ws.cell(row=row, column=col_idx, value="—").border = thin_border

            # Center all cells in row
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).alignment = center_align

        # Set column widths
        col_widths = [10, 8, 10, 10, 8, 6, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
